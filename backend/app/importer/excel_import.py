"""Migrate the client Excel master sheets into structured product records.

Reads each workbook twice — once for formulas (to derive the engine parameters)
and once for cached computed values (the migration "ground truth" snapshot). For
every product row it:

  * extracts identity fields (name, description, model, link, qty, category);
  * derives the engine parameters (loading_factor, client_markup, markup_base,
    list_uplift) by parsing the P / R / E formula multipliers;
  * captures the cached cost (P), client price (R), list (E) and total (H);
  * flags rows whose formulas deviate from the simple multiplicative pattern
    (hardcoded prices, odd multipliers) as manual overrides.

It deliberately does NOT touch the database — it returns plain dataclasses so it
can be unit-/parity-tested without MSSQL. The DB loader (Phase 2) consumes these.
"""

from __future__ import annotations

import glob
import os
import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from app.core.pricing import (
    PricingInputs, MarkupBase, compute_unit,
    DEFAULT_LIST_UPLIFT,
)

# Default location of the client workbooks (repo-root/working excel sheet from clients).
DEFAULT_SHEETS_DIR = os.environ.get(
    "EVAVO_SHEETS_DIR",
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "..",
                 "working excel sheet from clients"),
)

# How a (file, sheet) maps to a product category. Falls back to the sheet title.
CATEGORY_BY_SHEET = {
    "Salon Equipments": "Salon Equipment",
    "SALON MASTER QUOTE": "Salon Equipment",
    "Massage Beds": "Massage Beds",
    "MASSAGE BEDS": "Massage Beds",
    "STATIONARY BEDS": "Massage Beds",
    "ELECTRIC BED": "Massage Beds",
    "PORTABLE BEDS": "Massage Beds",
    "AYURVED MASSAGE BEDS": "Massage Beds",
    "Loungers": "Loungers",
    "ACCESSORIES": "Accessories",
}

# Column letters (1-based via openpyxl); the layout is consistent across sheets.
COL = {
    "sr": "A", "spec": "B", "product": "C", "desc": "D",
    "list_price": "E", "disc_unit": "F", "qty": "G", "total": "H",
    "link": "I", "model": "J", "client": "R",  # R = "Quote 2 Client" on most sheets
    "final_c2e": "P",  # P = "Final C2E" = true unit cost
}

_SIMPLE_MULT = re.compile(r"^=([A-Z]+)(\d+)\*([0-9]*\.?[0-9]+)$")          # =N4*1.5
_UPLIFT = re.compile(r"^=([A-Z]+)(\d+)\*\(1\+([0-9]+)%\)$")               # =F4*(1+10%)


def _cell(ws, col_letter: str, row: int):
    return ws.cell(row=row, column=column_index_from_string(col_letter)).value


def _num(v) -> Optional[float]:
    try:
        if v is None:
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


@dataclass
class ProductRecord:
    # identity
    name: str
    description: str
    model_no: str
    category: str
    product_link: str
    source_file: str
    source_sheet: str
    source_row: int
    default_qty: float

    # engine parameters (forward-compute -> migrated values)
    source_price_inr: float                 # c2e_inr basis stored in INR
    loading_factor: float
    client_markup: float
    markup_base: MarkupBase
    list_uplift: float

    # migrated cached snapshot (ground truth from the workbook)
    migrated_final_c2e: Optional[float]     # cached P (unit cost) — CONFIDENTIAL
    migrated_client_unit: Optional[float]   # cached R (unit selling price)
    migrated_list_price: Optional[float]    # cached E
    migrated_total: Optional[float]         # cached H

    # provenance / quality
    is_manual_override: bool = False
    override_reasons: list[str] = field(default_factory=list)

    def to_pricing_inputs(self) -> PricingInputs:
        return PricingInputs(
            source_price=self.source_price_inr,
            fx_rate=1.0,
            conversion_factor=1.0,
            loading_factor=self.loading_factor,
            client_markup=self.client_markup,
            list_uplift=self.list_uplift,
            markup_base=self.markup_base,
        )


def _category_for(sheet_title: str) -> str:
    key = sheet_title.strip().upper()
    for k, v in CATEGORY_BY_SHEET.items():
        if k.upper() == key:
            return v
    return sheet_title.strip().title()


def _find_header_row(ws) -> Optional[int]:
    for r in range(1, 13):
        v = _cell(ws, COL["sr"], r)
        if isinstance(v, str) and v.strip().upper().startswith("SR"):
            return r
    return None


def _parse_sheet(ws_f, ws_v, source_file: str) -> list[ProductRecord]:
    """ws_f = formula view, ws_v = cached-value view (same sheet)."""
    records: list[ProductRecord] = []
    hr = _find_header_row(ws_f)
    if hr is None:
        return records
    category = _category_for(ws_f.title)

    for r in range(hr + 1, ws_f.max_row + 1):
        name = _cell(ws_v, COL["spec"], r) or _cell(ws_v, COL["product"], r)
        client_cached = _num(_cell(ws_v, COL["client"], r))
        cost_cached = _num(_cell(ws_v, COL["final_c2e"], r))
        # A data row needs at least a name and a computed client price.
        if not name or client_cached is None:
            continue

        reasons: list[str] = []

        # --- loading_factor & the c2e base cell, from the P formula ---
        p_formula = _cell(ws_f, COL["final_c2e"], r)
        loading_factor = None
        c2e_base_letter = None
        m = _SIMPLE_MULT.match(str(p_formula or ""))
        if m:
            c2e_base_letter = m.group(1)
            loading_factor = float(m.group(3))
        else:
            reasons.append(f"P formula non-standard: {p_formula!r}")

        # --- client_markup & markup_base, from the R formula ---
        r_formula = _cell(ws_f, COL["client"], r)
        client_markup = None
        markup_base = MarkupBase.FINAL_C2E
        mr = _SIMPLE_MULT.match(str(r_formula or ""))
        if mr:
            r_base_letter = mr.group(1)
            client_markup = float(mr.group(3))
            # If R multiplies the same cell P multiplies (the c2e cell), the
            # markup is applied pre-loading (Accessories sheet). If it multiplies
            # P itself, it's applied to final_c2e (most sheets).
            if c2e_base_letter and r_base_letter == c2e_base_letter:
                markup_base = MarkupBase.C2E_INR
            elif r_base_letter == COL["final_c2e"]:
                markup_base = MarkupBase.FINAL_C2E
            else:
                markup_base = MarkupBase.FINAL_C2E
                reasons.append(f"R base {r_base_letter} unexpected")
        else:
            reasons.append(f"R formula non-standard: {r_formula!r}")

        # --- list_uplift, from the E formula (=F*(1+10%)) ---
        e_formula = _cell(ws_f, COL["list_price"], r)
        me = _UPLIFT.match(str(e_formula or ""))
        list_uplift = float(me.group(3)) / 100.0 if me else DEFAULT_LIST_UPLIFT

        # --- c2e_inr basis (cached value of the cell P references) ---
        c2e_inr = None
        if c2e_base_letter:
            c2e_inr = _num(_cell(ws_v, c2e_base_letter, r))
        if c2e_inr is None and cost_cached is not None and loading_factor:
            # Recover the basis from cached cost so the engine still reproduces it.
            c2e_inr = cost_cached / loading_factor
            reasons.append("c2e basis recovered from cached cost")

        # Fallbacks that still guarantee a usable record.
        if loading_factor is None and cost_cached is not None and c2e_inr:
            loading_factor = cost_cached / c2e_inr
        if loading_factor is None:
            loading_factor = 1.0
        if c2e_inr is None:
            c2e_inr = cost_cached if cost_cached is not None else client_cached
        if client_markup is None and cost_cached:
            client_markup = client_cached / cost_cached
        if client_markup is None:
            client_markup = client_cached / c2e_inr if c2e_inr else 1.0

        rec = ProductRecord(
            name=str(name).strip(),
            description=str(_cell(ws_v, COL["desc"], r) or "").strip(),
            model_no=str(_cell(ws_v, COL["model"], r) or "").strip(),
            category=category,
            product_link=str(_cell(ws_v, COL["link"], r) or "").strip(),
            source_file=os.path.basename(source_file),
            source_sheet=ws_f.title,
            source_row=r,
            default_qty=_num(_cell(ws_v, COL["qty"], r)) or 1.0,
            source_price_inr=c2e_inr,
            loading_factor=loading_factor,
            client_markup=client_markup,
            markup_base=markup_base,
            list_uplift=list_uplift,
            migrated_final_c2e=cost_cached,
            migrated_client_unit=client_cached,
            migrated_list_price=_num(_cell(ws_v, COL["list_price"], r)),
            migrated_total=_num(_cell(ws_v, COL["total"], r)),
            is_manual_override=bool(reasons),
            override_reasons=reasons,
        )
        records.append(rec)
    return records


def import_workbook(path: str) -> list[ProductRecord]:
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    out: list[ProductRecord] = []
    for ws_f in wb_f.worksheets:
        ws_v = wb_v[ws_f.title]
        out.extend(_parse_sheet(ws_f, ws_v, path))
    return out


def import_all(sheets_dir: str = DEFAULT_SHEETS_DIR) -> list[ProductRecord]:
    paths = sorted(glob.glob(os.path.join(sheets_dir, "*.xlsx")))
    out: list[ProductRecord] = []
    for p in paths:
        if os.path.basename(p).startswith("~$"):
            continue
        out.extend(import_workbook(p))
    return out
